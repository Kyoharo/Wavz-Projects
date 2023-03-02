#Truesigt_export 
from selenium  import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from time import sleep
import getpass
import glob
import os
import csv 
import openpyxl
import re






class InstaBot:
    def __init__(self,username,password):  
        #webdriver
        serv_obj = service= Service('geckodriver.exe')
        ops=webdriver.FirefoxOptions()
        #ps.headless=True        
        ops.add_argument("--disable-notifications")
        self.driver = webdriver.Firefox(service=serv_obj,options=ops)
        #launch
        self.driver.get("http://web-tier.egyptpost.local:8445")
        self.driver.implicitly_wait(10)
        self.username= username
        self.password= password
        #username
        self.driver.find_element(By.XPATH, "//input[@id='user_login']").send_keys(username)
        #password
        self.driver.find_element(By.XPATH, "//input[@id='login_user_password']").send_keys(password)
        #login
        self.driver.find_element(By.XPATH, "//button[@id='login-jsp-btn']").click()
        self.driver.maximize_window()


    def choose_filter(self):
        #console
        self.driver.find_element(By.XPATH, "//a[normalize-space()='Console']").click()
        #Monitoring
        self.driver.find_element(By.XPATH, "//a[normalize-space()='Ticket Console']").click()
       


    def get_tickets_data(self):
         #select all 
        tickets = self.driver.find_elements(By.XPATH, "//div[contains(@class,'ngCanvas')]/child::div")
        for row in range(1,len(tickets)+1):
            ticket= self.driver.find_element(By.XPATH, "//div[contains(@class,'ngCanvas')]/child::div["+str(row)+"]")
            ticket.click()  
            #Excel 
            sheet_path = 'ff.xlsx'
            wb = openpyxl.load_workbook(sheet_path)
            sheet = wb['Sheet1']     
            #get_Incident_1
            ele_incident = self.driver.find_element(By.XPATH, "//div[@ux-id='field_id']").text
            sheet.cell(row,1).value = ele_incident
            wb.save(sheet_path)
            #get_ip_2
            ele_title_bar = self.driver.find_element(By.XPATH, "//div[@class='title-bar__summary ng-binding ng-scope']").text
            #Regular Expresstion to get the ip
            Reg_ip = re.compile(r"\d+\.\d+\.\d+\.\d+")
            convert_ip = Reg_ip.findall(ele_title_bar)
            ip1= "".join(convert_ip)
            sheet.cell(row,2).value = ip1
            #----------
             #Regular Expresstion to get the AArea
            Reg_ip = re.compile(r"[\u0600-\u06FF ]+[\u0600-\u06FF ]+")
            convert_ip = Reg_ip.findall(ele_title_bar)
            #convert the_area_3
            area= "".join(convert_ip)
            sheet.cell(row,3).value = area
            wb.save(sheet_path)
            try:
                #get_vednor_4
                ele_vendor = self.driver.find_element(By.XPATH, "//div[@get-field-value='getFieldValue']//span").text
                sheet.cell(row,4).value = ele_vendor
                wb.save(sheet_path)
                #get_Vendor Ticket Number_5
                ele_vendor_Number = self.driver.find_element(By.XPATH, "//div[@ux-id='field_vendorTicketNumber']//span").text
                sheet.cell(row,5).value = ele_vendor_Number
                wb.save(sheet_path)
            except:
                pass
            #get_Submit Date_6
            Submit_Dater = self.driver.find_element(By.XPATH, "//div[@class='custom-field col-md-12']//div[@class='clearfix custom-field__date_time_container ng-isolate-scope']//span").text
            sheet.cell(row,6).value = Submit_Dater
            wb.save(sheet_path)
            self.driver.back()


    def Quit(self):
        self.driver.quit()





        

          

# user_id=input('Enter Your Username:')
# password=getpass('Enter Password:')
# #user_id=""
#password=""

mybot=InstaBot("w_talaat", "Talaat@999")
mybot.choose_filter()
mybot.get_tickets_data()
mybot.Quit()


