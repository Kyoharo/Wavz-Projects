#excel modules 
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import re



import tkinter as tk
from tkinter import filedialog

# Create the main window
root = tk.Tk()
root.geometry("600x500")

# Add the background image
bg_image = tk.PhotoImage(file=r"C:\Users\abdoa\Desktop\wavz.png")
bg_label = tk.Label(root, image=bg_image)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# Create the "Browse" buttons
file1_path = ""
file2_path = ""
file3_path = ""

def browse_file1():
    global file1_path
    file1_path = filedialog.askopenfilename()
    # Do something with the file (e.g. display it)

def browse_file2():
    global file2_path
    file2_path = filedialog.askopenfilename()
    # Do something with the file (e.g. display it)

def browse_file3():
    global file3_path
    file3_path = filedialog.askopenfilename()
    # Do something with the file (e.g. display it)

browse_button1 = tk.Button(root, text="Alarm", command=browse_file1, height = 3, width = 10)
browse_button1.place(relx=0.2, rely=0.2, anchor='center')
browse_button2 = tk.Button(root, text="Reference", command=browse_file2, height = 3, width = 10)
browse_button2.place(relx=0.5, rely=0.2, anchor='center')
browse_button3 = tk.Button(root, text="Result", command=browse_file3, height = 3, width = 10)
browse_button3.place(relx=0.8, rely=0.2, anchor='center')

# Create the "Start" button
def start_app():
    print("File 1: ", file1_path)
    print("File 2: ", file2_path)
    print("File 3: ", file3_path)
    # Insert your code here to execute the rest of the code



# Run the main loop
root.mainloop()



#for print
#------------------
import sys
import time
def delay_print(s):
    for c in s:
        sys.stdout.write(c)
        sys.stdout.flush()
        time.sleep(0.01)






wb = openpyxl.load_workbook(file1_path) 
#pick tap sheet 
ws = wb['Sheet1']                           #to pick the sheet u want to work on
#get max row of check sheet
row = ws.max_row+1                                   # to add max row to variable          (ws sheet)
#-----------------------


#open the reverance sheet (2)xx
ree = openpyxl.load_workbook(file2_path) 
#pick tap sheet 
rev = ree['Sheet1'] 
#get max row of check sheet
row2 = rev.max_row+1                                   # to add max row to variable         (rev  sheet)
#-----------------------------------



#Open Result sheet (3)
result_sheet = openpyxl.load_workbook(file3_path) 
res_sheet = result_sheet['Sheet1'] 
row3 = res_sheet.max_row+1                                   # to add max row to variable         (rev  sheet)
#conter for Result sheet 
ii =2

MY_ip_list = []
for i in range(1,row):
    #column of ip address from check sheet (1)
        ip = ws.cell(i,5).value
        port = ws.cell(i,4).value   
        try:

            #----------------------------------
            match_issue = 'Ce0|ATM|Vl|Tu'
            Reg_char = re.compile(match_issue)
            convert_char = Reg_char.findall(ip)
            char = "".join(convert_char)


            #get ip format by regular excperation 
            Reg_ip = re.compile(r"\d+\.\d+\.\d+\.\d+")
            convert_ip = Reg_ip.findall(ip)
            #convert the ip to string (1)
            ip1= "".join(convert_ip)
            #-----------------

            #second sheet (reverance) (2)
            for j in range(1,row2):
                #column of ip address (2)
                ipp = rev.cell(j,1).value
                try:
                    #get ip format 
                    Reg_ip = re.compile(r"\d+\.\d+\.\d+\.\d+")
                    conver_ip2 = Reg_ip.findall(ipp)
                    #convert the ip to string (2)
                    ip2= "".join(conver_ip2)
                    #-----------------
                    #check ip of (check sheet(1)) is the same of (reverance sheet(2))
                    if ip1 == ip2 and ip1 != "":
                        if ip1 not in MY_ip_list and port == "Port Status Problem" and char != "":
                        #get information from referance sheet
                            MY_ip_list.append(ip1)
                            area = rev.cell(j,2).value
                            Region = rev.cell(j,4).value
                            wan_ip = rev.cell(j,32).value
                            tunnel_ip = rev.cell(j,33).value
                            network_id = rev.cell(j,23).value
                            line_type_id = rev.cell(j,24).value
                            main_order_id = rev.cell(j,18).value
                            backup_order_id = rev.cell(j,19).value
                            project_id = rev.cell(j,20).value


                            #ASSIGN 
                            #write value of ip-area- region
                            res_sheet.cell(ii,1).value = ip1                #ip 
                            res_sheet.cell(ii,2).value = ip                 #ip with description 
                            res_sheet.cell(ii,3).value = wan_ip             # WAN IP
                            res_sheet.cell(ii,4).value = tunnel_ip          #tunnel_ip
                            res_sheet.cell(ii,5).value = port               #Alarm
                            res_sheet.cell(ii,7).value = Region             # Reagion
                            res_sheet.cell(ii,6).value = area               # area
                            res_sheet.cell(ii,8).value = network_id         #network_id
                            res_sheet.cell(ii,9).value = line_type_id       #line_type_id
                            res_sheet.cell(ii,10).value = main_order_id     #main_order_id
                            res_sheet.cell(ii,11).value = backup_order_id   #backup_order_id
                            res_sheet.cell(ii,12).value = project_id        #project_id
                            #save the sheet
                            result_sheet.save(path_3)
                            ii+=1 
                        else:
                            pass

                except Exception:
                    pass
            if ip1 not in MY_ip_list and port =="Port Status Problem" and char != "":  
                MY_ip_list.append(ip1)
                res_sheet.cell(ii,1).value = ip1 
                res_sheet.cell(ii,2).value = ip
                res_sheet.cell(ii,5).value = port
                res_sheet.cell(ii,6).value = "not match"
                res_sheet.cell(ii,7).value = "not match"
                
                result_sheet.save(path_3)
                ii+=1

        except Exception:
            pass

print(MY_ip_list)


    

