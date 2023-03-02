# -*- coding: utf-8 -*-
"""
Created on Sun Dec 11 17:10:32 2022

@author: abdoa
"""

# -*- coding: utf-8 -*-
"""
Created on Sun Dec 11 12:25:52 2022

@author: abdoa
"""
#excel modules 
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
#--------------
#operating system module
import os
#--------------
#time module
import time
from datetime import date
from datetime import date
#--------------
#Selenium modules 
from selenium  import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#--------------
def launch(url):
    global browser 
    browser = webdriver.Firefox(service= Service(r'\\10.199.199.35\soc team\Abdelrahman Ataa\Driver\geckodriver.exe'))
    browser.get(url)
    #Username
    link = WebDriverWait(browser,10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#LS_username_input"))                                           )
    link.click()
    link.send_keys("W_abdelrahman.ataa")
    time.sleep(2)
    #password
    link = browser.find_element(By.CSS_SELECTOR, "#LS_password_input")
    link.click()
    link.send_keys("a1591997A!")
    time.sleep(3)
    #login
    link = browser.find_element(By.CSS_SELECTOR, "#LS_login_button")
    link.click()
    time.sleep(10)

def services(service_by_css_selector):
    #Men
    element = WebDriverWait(browser,5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#MNV_dashboards_button > svg:nth-child(1)"))                                           )
    element.click()
    #call center
    time.sleep(2)
    element = WebDriverWait(browser,5).until(EC.presence_of_element_located((By.CSS_SELECTOR, service_by_css_selector))                                       )
    element.click()
    






def get_info(area_by_xpath,area_issue_by_xpath):
    #get Service name 
    global MyRow
    #get الجيزة issues 
    element = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.XPATH, area_issue_by_xpath))                                       )
    ele4 = element.text
    if ele4 == '0':
        print("no issues ")
    else:
        wb = openpyxl.load_workbook(r'\\10.199.199.35\soc team\SOC_Daily Report\2023\Feb\Nexthink feb Report.xlsx') 
        ws = wb['Sheet1']                           #to pick the sheet u want to work on
        #get the row
        #find service name
        element = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.ID, "DNV_dashboard_span"))                                       )
        ele1 = element.text
        row = ws.max_row +1
        for j in range(MyRow,row):
            if ws.cell(j, 1).value == None:
                global empty_Row
                empty_Row = j
                break

        print(empty_Row)
        #----------
        #get time
        element = browser.find_element(By.ID, "TNV_scope_text")
        ele2 = element.text

        #---------
        #---------
        #get الجيزة
        element = browser.find_element(By.XPATH, area_by_xpath)
        ele3 = element.text

        #get issue area
        element = browser.find_element(By.XPATH, area_issue_by_xpath)
        ele4 = element.text
        #---------
        

        #**************************************** Save ***************************************
        #find service name
        ws.cell(empty_Row,1).value = ele1
        #get time
        ws.cell(empty_Row,3).value = ele2
        #get الجيزة
        ws.cell(empty_Row,5).value = ele3
        #get الجيزة issues 
        ws.cell(empty_Row,6).value = ele4
        
        #------------------------------------
        #Date
        from datetime import date
        today = date.today()
        d1 = today.strftime("%d-%m-%Y")
        result = ''.join(d1)
        ws.cell(empty_Row,4).value = result
        
        #---------------
        #Save
        wb.save(r'\\10.199.199.35\soc team\SOC_Daily Report\2023\Feb\Nexthink feb Report.xlsx') 





#Start launch

try: 
    launch("https://nxportal.egyptpost.local")
except:
    browser.refresh()
    launch("https://nxportal.egyptpost.local")


#import in Excel
wb = openpyxl.load_workbook(r'\\10.199.199.35\soc team\SOC_Daily Report\2023\Feb\Nexthink feb Report.xlsx') 
sheets = wb.sheetnames
ws = wb['Sheet1']                           #to pick the sheet u want to work on
row = ws.max_row +1
#get the row
for i in range(1,row):                                         
    if ws.cell(i, 1).value == None:
        MyRow = i
        break

#-------------------------
#Get Date
from datetime import date
today = date.today()
d1 = today.strftime("%d-%m-%Y")



def main_services(service_by_css):
    #call center
    services(service_by_css)
    #find service name
    element = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.ID, "DNV_dashboard_span"))                                       )
    global service_by_name
    service_by_name = element.text
    
    if service_by_name != service_name:
        services("#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_84424398-eb97-49b8-b862-78cf2c008538_1")
        time.sleep(3)
        services(service_by_css_for_name)

    element = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[5]/div[2]/div/div[1]/div/div[1]/div[1]/div/div[1]/div[1]/span[2]"))                                           )
    check_status = element.text
    if check_status == '0' :
        print("service status: " + check_status)
    else:
        #-

        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[2]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[2]/td[2]")
        #الجيزة
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[3]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[3]/td[2]")
        #القاهرة
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[4]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[4]/td[2]")
        #المباني الادارية
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[5]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[5]/td[2]")
        #شرق الدلتا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[6]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[6]/td[2]")
        #شمال الدلتا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[7]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[7]/td[2]")
        #غرب الدلتا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[8]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[8]/td[2]")
        #مصر العليا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[9]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[9]/td[2]")
        #مصر الوسطا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[10]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[10]/td[2]")
        #وسط الدلتا
        get_info("/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[11]/td[1]","/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[6]/div[6]/div[2]/div/div[2]/div[1]/div/div[3]/div[4]/div[1]/div/table/tbody/tr[11]/td[2]")
        #------------------




        


#call Center
service_by_css_for_name = "#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_84424398-eb97-49b8-b862-78cf2c008538_1"
service_name = "Call Center"
main_services(service_by_css_for_name)

#E-Counter
service_by_css_for_name = "#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_28773179-8532-4e7e-a3b4-4b1b74dd8418_1"
service_name = "E-Counter"
main_services(service_by_css_for_name)

#Postal Back office
service_by_css_for_name = "#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_4db2c7fe-a30d-44ea-9c50-75c6ba58d414_1"
service_name = "Postal Back Office"
main_services(service_by_css_for_name)

#Postal Front office
service_by_css_for_name ="#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_060b57e6-04ee-48d8-8dd6-57253c08ec15_1"
service_name = "Postal Front Office"
main_services(service_by_css_for_name)

#SAP-Services 
service_by_css_for_name ="#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_2a9d6b58-14e7-47b4-aa31-2fd2e2537074_1"
service_name = "SAP_Service"
main_services(service_by_css_for_name)

#حوالات فورية
service_by_css_for_name = "#MNV_dashboard_item_589f0156-70b6-4dba-bf24-2afe56e11bff_bdd87d50-7bb6-4767-b7ed-0946b0060e90_1"
service_name = "حوالات فورية"
main_services(service_by_css_for_name)

browser.quit()



